#!/usr/bin/env python3
"""
تولید نمودارهای فصل ۴ — قابل ویرایش

نحوه استفاده:
  1. داده‌ها را در docs/chapter4-data.json ویرایش کنید
  2. اجرا: python3 docs/generate_chapter4_figures.py
  3. خروجی:
     - PNG  → docs/chapter4-figures/*.png   (برای درج در Word)
     - SVG  → docs/chapter4-figures/*.svg   (قابل ویرایش در Inkscape / Illustrator / Word)
     - Excel → docs/chapter4-figures/chapter4-data.xlsx
"""

from __future__ import annotations

import json
from pathlib import Path

import arabic_reshaper
import matplotlib.font_manager as fm
import matplotlib.pyplot as plt
import numpy as np
from bidi.algorithm import get_display
from matplotlib.ticker import FuncFormatter

BASE_DIR = Path(__file__).parent
DATA_FILE = BASE_DIR / "chapter4-data.json"
OUTPUT_DIR = BASE_DIR / "chapter4-figures"
FONT_PATH = BASE_DIR / "fonts" / "Vazirmatn-Regular.ttf"

fm.fontManager.addfont(str(FONT_PATH))
PERSIAN_FP = fm.FontProperties(fname=str(FONT_PATH), size=11)
PERSIAN_FP_SM = fm.FontProperties(fname=str(FONT_PATH), size=10)
PERSIAN_FP_TITLE = fm.FontProperties(fname=str(FONT_PATH), size=13)
LATIN_FP = fm.FontProperties(family="DejaVu Sans", size=10)
LATIN_FP_SM = fm.FontProperties(family="DejaVu Sans", size=9)

plt.rcParams.update(
    {
        "font.family": "DejaVu Sans",
        "axes.unicode_minus": False,
        "figure.dpi": 150,
        "savefig.dpi": 300,
        "savefig.bbox": "tight",
        "svg.fonttype": "none",
    }
)

_PERSIAN_DIGITS = str.maketrans("0123456789.", "۰۱۲۳۴۵۶۷۸۹٫")

COLORS = {
    "primary": "#2E86AB",
    "secondary": "#A23B72",
    "accent": "#F18F01",
    "success": "#3A7D44",
    "neutral": "#6C757D",
}

DATA: dict = {}
USE_FA_DIGITS = True


def load_data() -> dict:
    with open(DATA_FILE, encoding="utf-8") as f:
        return json.load(f)


def fa(text: str) -> str:
    if not text:
        return text
    return get_display(arabic_reshaper.reshape(text))


def fa_num(value, decimals: int | None = None, percent: bool = False) -> str:
    if not USE_FA_DIGITS:
        s = f"{float(value):.{decimals}f}" if decimals is not None else str(value)
        return f"{s}%" if percent else s
    if decimals is not None:
        formatted = f"{float(value):.{decimals}f}"
    else:
        formatted = str(value)
    translated = formatted.translate(_PERSIAN_DIGITS)
    return f"{translated}٪" if percent else translated


def fa_labels(labels: list[str]) -> list[str]:
    return [fa(label) for label in labels]


def set_persian_title(fig, text: str) -> None:
    fig.suptitle(fa(text), fontproperties=PERSIAN_FP_TITLE, y=1.02)


def set_persian_ylabel(ax, text: str) -> None:
    ax.set_ylabel(fa(text), fontproperties=PERSIAN_FP)


def set_persian_xlabel(ax, text: str) -> None:
    ax.set_xlabel(fa(text), fontproperties=PERSIAN_FP)


def set_latin_xticklabels(ax, labels: list[str], rotation: int = 0, ha: str = "center") -> None:
    ax.set_xticklabels(labels, fontproperties=LATIN_FP, rotation=rotation, ha=ha)


def set_persian_xticklabels(ax, labels: list[str], rotation: int = 0, ha: str = "center") -> None:
    ax.set_xticklabels(fa_labels(labels), fontproperties=PERSIAN_FP_SM, rotation=rotation, ha=ha)


def set_persian_yticklabels(ax, labels: list[str]) -> None:
    ax.set_yticklabels(fa_labels(labels), fontproperties=PERSIAN_FP_SM)


def persian_yformatter(decimals: int = 0):
    return FuncFormatter(lambda x, _pos: fa_num(x, decimals=decimals) if decimals else fa_num(int(x)))


def save_fig(fig, stem: str) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    settings = DATA.get("settings", {})
    if settings.get("export_png", True):
        fig.savefig(OUTPUT_DIR / f"{stem}.png", format="png", facecolor="white", edgecolor="none")
        print(f"  PNG: {stem}.png")
    if settings.get("export_svg", True):
        fig.savefig(OUTPUT_DIR / f"{stem}.svg", format="svg", facecolor="white", edgecolor="none")
        print(f"  SVG: {stem}.svg")
    plt.close(fig)


def tech_labels() -> list[str]:
    return DATA["lpwan_technologies"]["labels"]


# ── نمودارها ──────────────────────────────────────────────────────────────


def fig_clustering_inertia_silhouette():
    d = DATA["clustering_evaluation"]
    k = np.array(d["k"])
    inertia = np.array(d["inertia"])
    silhouette = np.array(d["silhouette"])

    fig, ax1 = plt.subplots(figsize=(8, 5))
    ax1.plot(k, inertia, "o-", color=COLORS["primary"], linewidth=2, markersize=8, label="Inertia (WCSS)")
    set_persian_xlabel(ax1, "تعداد خوشه‌ها (k)")
    ax1.set_ylabel("Inertia (WCSS)", color=COLORS["primary"], fontproperties=LATIN_FP)
    ax1.set_xticks(k)
    ax1.set_xticklabels([fa_num(t) for t in k], fontproperties=PERSIAN_FP_SM)
    ax1.yaxis.set_major_formatter(persian_yformatter(1))
    ax1.grid(True, alpha=0.3, linestyle="--")

    ax2 = ax1.twinx()
    ax2.plot(k, silhouette, "s-", color=COLORS["accent"], linewidth=2, markersize=8, label=fa("ضریب Silhouette"))
    ax2.set_ylabel(fa("ضریب Silhouette"), color=COLORS["accent"], fontproperties=PERSIAN_FP)
    ax2.yaxis.set_major_formatter(FuncFormatter(lambda x, _pos: fa_num(x, decimals=3)))

    sel = d["selected_k"]
    ax1.axvline(x=sel, color=COLORS["success"], linestyle="--", alpha=0.7, linewidth=1.5)
    ax1.annotate(
        fa(f"k={sel} (انتخاب نهایی)"),
        xy=(sel, inertia[list(d["k"]).index(sel)]),
        xytext=(sel - 1.2, max(inertia) * 0.55),
        fontproperties=PERSIAN_FP_SM,
        color=COLORS["success"],
        arrowprops=dict(arrowstyle="->", color=COLORS["success"], lw=1.2),
    )
    ax1.legend(loc="upper right", prop=LATIN_FP)
    set_persian_title(fig, d["title"])
    save_fig(fig, "fig4-1_inertia_silhouette_vs_k")


def fig_clustering_inertia_only():
    d = DATA["clustering_evaluation"]
    k, inertia = d["k"], d["inertia"]
    fig, ax = plt.subplots(figsize=(7, 5))
    bars = ax.bar(k, inertia, color=COLORS["primary"], alpha=0.85, edgecolor="white", width=0.6)
    ax.plot(k, inertia, "o--", color=COLORS["secondary"], linewidth=1.5, markersize=6)
    bars[d["k"].index(d["selected_k"])].set_color(COLORS["success"])
    set_persian_xlabel(ax, "تعداد خوشه‌ها (k)")
    ax.set_ylabel("Inertia (WCSS)", fontproperties=LATIN_FP)
    ax.set_xticks(k)
    ax.set_xticklabels([fa_num(t) for t in k], fontproperties=PERSIAN_FP_SM)
    ax.yaxis.set_major_formatter(persian_yformatter(1))
    ax.grid(True, axis="y", alpha=0.3, linestyle="--")
    set_persian_title(fig, "نمودار ۴-۲ — تغییرات Inertia در مقادیر مختلف k")
    save_fig(fig, "fig4-2_inertia_vs_k")


def fig_clustering_silhouette_only():
    d = DATA["clustering_evaluation"]
    k, silhouette = d["k"], d["silhouette"]
    fig, ax = plt.subplots(figsize=(7, 5))
    bars = ax.bar(k, silhouette, color=COLORS["accent"], alpha=0.85, edgecolor="white", width=0.6)
    ax.plot(k, silhouette, "o--", color=COLORS["secondary"], linewidth=1.5, markersize=6)
    bars[d["k"].index(d["selected_k"])].set_color(COLORS["success"])
    set_persian_xlabel(ax, "تعداد خوشه‌ها (k)")
    set_persian_ylabel(ax, "ضریب Silhouette")
    ax.set_xticks(k)
    ax.set_xticklabels([fa_num(t) for t in k], fontproperties=PERSIAN_FP_SM)
    ax.yaxis.set_major_formatter(FuncFormatter(lambda x, _pos: fa_num(x, decimals=3)))
    ax.set_ylim(0.28, 0.52)
    ax.grid(True, axis="y", alpha=0.3, linestyle="--")
    set_persian_title(fig, "نمودار ۴-۳ — تغییرات ضریب Silhouette در مقادیر مختلف k")
    save_fig(fig, "fig4-3_silhouette_vs_k")


def fig_cluster_sizes():
    d = DATA["cluster_sizes"]
    labels = fa_labels(d["labels"])
    fig, ax = plt.subplots(figsize=(9, 5))
    bars = ax.bar(labels, d["sizes"], color=list(COLORS.values())[:5], edgecolor="white", width=0.65)
    for bar, size in zip(bars, d["sizes"]):
        ax.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height() + 0.05,
            fa_num(size),
            ha="center",
            va="bottom",
            fontproperties=PERSIAN_FP,
            fontweight="bold",
        )
    set_persian_ylabel(ax, "تعداد اعضا")
    ax.set_ylim(0, 4)
    ax.set_yticks([0, 1, 2, 3, 4])
    ax.set_yticklabels([fa_num(i) for i in range(5)], fontproperties=PERSIAN_FP_SM)
    ax.grid(True, axis="y", alpha=0.3, linestyle="--")
    set_persian_title(fig, d["title"])
    save_fig(fig, "fig4-4_cluster_sizes")


def fig_ahp_weights():
    d = DATA["ahp_weights"]
    criteria = fa_labels(d["criteria"])
    fig, ax = plt.subplots(figsize=(9, 6))
    y_pos = np.arange(len(criteria))
    bars = ax.barh(y_pos, d["percent"], color=COLORS["primary"], height=0.6, edgecolor="white")
    ax.set_yticks(y_pos)
    ax.set_yticklabels(criteria, fontproperties=PERSIAN_FP_SM)
    set_persian_xlabel(ax, "سهم (%)")
    ax.set_xlim(0, 20)
    ax.xaxis.set_major_formatter(persian_yformatter(0))
    for bar, w in zip(bars, d["percent"]):
        ax.text(
            bar.get_width() + 0.3,
            bar.get_y() + bar.get_height() / 2,
            fa_num(w, decimals=2, percent=True),
            va="center",
            fontproperties=PERSIAN_FP_SM,
        )
    ax.grid(True, axis="x", alpha=0.3, linestyle="--")
    set_persian_title(fig, d["title"])
    save_fig(fig, "fig4-5_ahp_weights")


def fig_base_vs_adaptive_weights():
    d = DATA["weight_comparison"]
    criteria = fa_labels(d["criteria"])
    x = np.arange(len(criteria))
    width = 0.35
    fig, ax = plt.subplots(figsize=(11, 6))
    ax.bar(x - width / 2, d["base"], width, label=fa("وزن پایه AHP"), color=COLORS["primary"], alpha=0.85)
    ax.bar(x + width / 2, d["adaptive"], width, label=fa("وزن تطبیقی نهایی"), color=COLORS["accent"], alpha=0.85)
    ax.set_xticks(x)
    set_persian_xticklabels(ax, d["criteria"], rotation=35, ha="right")
    set_persian_ylabel(ax, "وزن معیار")
    ax.set_ylim(0, 0.45)
    ax.yaxis.set_major_formatter(FuncFormatter(lambda v, _pos: fa_num(v, decimals=2)))
    ax.legend(loc="upper left", prop=PERSIAN_FP_SM)
    ax.grid(True, axis="y", alpha=0.3, linestyle="--")
    set_persian_title(fig, d["title"])
    save_fig(fig, "fig4-6_base_vs_adaptive_weights")


def fig_adaptive_weights_only():
    d = DATA["adaptive_weights"]
    criteria = fa_labels(d["criteria"])
    pct = [w * 100 for w in d["weights"]]
    fig, ax = plt.subplots(figsize=(9, 6))
    y_pos = np.arange(len(criteria))
    colors = [COLORS["accent"] if w < 0.1 else COLORS["primary"] if w < 0.2 else COLORS["success"] for w in d["weights"]]
    bars = ax.barh(y_pos, pct, color=colors, height=0.6, edgecolor="white")
    ax.set_yticks(y_pos)
    ax.set_yticklabels(criteria, fontproperties=PERSIAN_FP_SM)
    set_persian_xlabel(ax, "سهم (%)")
    ax.xaxis.set_major_formatter(persian_yformatter(0))
    for bar, p in zip(bars, pct):
        ax.text(
            bar.get_width() + 0.5,
            bar.get_y() + bar.get_height() / 2,
            fa_num(p, decimals=1, percent=True),
            va="center",
            fontproperties=PERSIAN_FP_SM,
        )
    ax.grid(True, axis="x", alpha=0.3, linestyle="--")
    set_persian_title(fig, d["title"])
    save_fig(fig, "fig4-7_adaptive_weights")


def _bar_chart_tech(stem: str, title: str, values: list[float], decimals: int, ylabel: str, ylim_top: float):
    """نمودار ستونی فناوری‌ها — محور X عددی + برچسب لاتین جدا."""
    techs = tech_labels()
    x = np.arange(len(techs))
    colors = [COLORS["success"], COLORS["accent"], COLORS["secondary"]]

    fig, ax = plt.subplots(figsize=(7, 5))
    bars = ax.bar(x, values, color=colors, edgecolor="white", width=0.55)
    ax.set_xticks(x)
    set_latin_xticklabels(ax, techs)
    set_persian_ylabel(ax, ylabel)
    ax.set_ylim(0, ylim_top)
    ax.yaxis.set_major_formatter(FuncFormatter(lambda v, _pos: fa_num(v, decimals=decimals)))
    for bar, val in zip(bars, values):
        ax.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height() + ylim_top * 0.02,
            fa_num(val, decimals=decimals),
            ha="center",
            va="bottom",
            fontproperties=PERSIAN_FP_SM,
        )
    ax.grid(True, axis="y", alpha=0.3, linestyle="--")
    set_persian_title(fig, title)
    save_fig(fig, stem)


def fig_topsis_closeness():
    d = DATA["topsis"]
    _bar_chart_tech("fig4-8_topsis_closeness", d["title"], d["closeness"], 4, "ضریب نزدیکی (C_i)", 1.0)


def fig_vikor_q_only():
    d = DATA["vikor"]
    _bar_chart_tech("fig4-9b_vikor_q_index", d["title_q"], d["q"], 3, "شاخص نهایی Q_i", 1.15)


def fig_vikor_indices():
    d = DATA["vikor"]
    techs = tech_labels()
    x = np.arange(len(techs))
    width = 0.25
    fig, ax = plt.subplots(figsize=(8, 5))
    ax.bar(x - width, d["s"], width, label=fa("S_i (سودمندی گروهی)"), color=COLORS["primary"])
    ax.bar(x, d["r"], width, label=fa("R_i (نارضایتی فردی)"), color=COLORS["accent"])
    ax.bar(x + width, d["q"], width, label=fa("Q_i (شاخص نهایی)"), color=COLORS["secondary"])
    ax.set_xticks(x)
    set_latin_xticklabels(ax, techs)
    set_persian_ylabel(ax, "مقدار شاخص")
    ax.yaxis.set_major_formatter(FuncFormatter(lambda v, _pos: fa_num(v, decimals=3)))
    ax.legend(loc="upper left", prop=PERSIAN_FP_SM)
    ax.grid(True, axis="y", alpha=0.3, linestyle="--")
    set_persian_title(fig, d["title_indices"])
    save_fig(fig, "fig4-9_vikor_indices")


def fig_copras_results():
    d = DATA["copras"]
    techs = tech_labels()
    x = np.arange(len(techs))
    width = 0.35
    fig, ax1 = plt.subplots(figsize=(8, 5))
    ax1.bar(x - width / 2, d["q"], width, label=fa("Q_i (اهمیت نسبی)"), color=COLORS["primary"])
    ax1.set_ylabel("Q_i", fontproperties=LATIN_FP)
    ax1.set_xticks(x)
    set_latin_xticklabels(ax1, techs)
    ax1.yaxis.set_major_formatter(FuncFormatter(lambda v, _pos: fa_num(v, decimals=2)))
    ax1.grid(True, axis="y", alpha=0.3, linestyle="--")

    ax2 = ax1.twinx()
    ax2.bar(x + width / 2, d["n"], width, label=fa("N_i (درجه مطلوبیت %)"), color=COLORS["accent"], alpha=0.85)
    set_persian_ylabel(ax2, "N_i (%)")
    ax2.set_ylim(0, 115)
    ax2.yaxis.set_major_formatter(persian_yformatter(0))

    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(lines1 + lines2, labels1 + labels2, loc="upper right", prop=PERSIAN_FP_SM)
    set_persian_title(fig, d["title"])
    save_fig(fig, "fig4-10_copras_results")


def fig_ranking_comparison():
    d = DATA["ranking_comparison"]
    techs = tech_labels()
    x = np.arange(len(techs))
    width = 0.25
    fig, ax = plt.subplots(figsize=(8, 5))
    ax.bar(x - width, d["topsis"], width, label="TOPSIS", color=COLORS["primary"])
    ax.bar(x, d["vikor"], width, label="VIKOR", color=COLORS["accent"])
    ax.bar(x + width, d["copras"], width, label="COPRAS", color=COLORS["secondary"])
    ax.set_xticks(x)
    set_latin_xticklabels(ax, techs)
    set_persian_ylabel(ax, "رتبه")
    ax.set_yticks([1, 2, 3])
    ax.set_yticklabels([fa_num(i) for i in [1, 2, 3]], fontproperties=PERSIAN_FP_SM)
    ax.invert_yaxis()
    ax.legend(loc="lower right", prop=LATIN_FP)
    ax.grid(True, axis="y", alpha=0.3, linestyle="--")
    set_persian_title(fig, d["title"])
    save_fig(fig, "fig4-11_ranking_comparison")


def fig_spearman_correlation():
    d = DATA["spearman"]
    pairs = fa_labels(d["pairs"])
    fig, ax = plt.subplots(figsize=(7, 5))
    bars = ax.bar(pairs, d["rho"], color=COLORS["success"], edgecolor="white", width=0.5)
    for bar in bars:
        ax.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height() - 0.08,
            fa_num(1.0, decimals=4),
            ha="center",
            va="top",
            fontproperties=PERSIAN_FP,
            fontweight="bold",
            color="white",
        )
    set_persian_ylabel(ax, "ضریب همبستگی اسپیرمن (rho)")
    ax.set_ylim(0, 1.15)
    ax.yaxis.set_major_formatter(FuncFormatter(lambda v, _pos: fa_num(v, decimals=1)))
    ax.axhline(y=1.0, color=COLORS["neutral"], linestyle="--", alpha=0.5)
    ax.grid(True, axis="y", alpha=0.3, linestyle="--")
    set_persian_title(fig, d["title"])
    save_fig(fig, "fig4-12_spearman_correlation")


def export_excel() -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, LineChart, Reference
        from openpyxl.chart.label import DataLabelList
        from openpyxl.styles import Alignment, Font
    except ImportError:
        print("  Excel: skipped (openpyxl not installed)")
        return

    wb = Workbook()
    wb.remove(wb.active)
    techs = tech_labels()

    def clean_label(text: str) -> str:
        return text.replace("\n", " — ")

    def write_title(ws, title: str, cols: int = 6) -> int:
        ws.cell(row=1, column=1, value=title)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
        ws.cell(row=1, column=1).font = Font(bold=True, size=12)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", wrap_text=True)
        return 3

    def write_headers(ws, row: int, headers: list) -> None:
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)

    def write_rows(ws, start_row: int, rows: list) -> int:
        for ri, row in enumerate(rows, start_row):
            for ci, val in enumerate(row, 1):
                ws.cell(row=ri, column=ci, value=val)
        return start_row + len(rows) - 1

    def place_chart(ws, chart, anchor: str) -> None:
        chart.width = 20
        chart.height = 12
        ws.add_chart(chart, anchor)

    def col_chart(ws, header_row: int, last_row: int, title: str, y_title: str = "", data_cols: int = 2):
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "clustered"
        chart.style = 10
        chart.title = title
        if y_title:
            chart.y_axis.title = y_title
        data = Reference(ws, min_col=2, min_row=header_row, max_col=1 + data_cols, max_row=last_row)
        cats = Reference(ws, min_col=1, min_row=header_row + 1, max_row=last_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        return chart

    def bar_h_chart(ws, header_row: int, last_row: int, title: str, y_title: str = ""):
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = title
        if y_title:
            chart.y_axis.title = y_title
        data = Reference(ws, min_col=2, min_row=header_row, max_col=2, max_row=last_row)
        cats = Reference(ws, min_col=1, min_row=header_row + 1, max_row=last_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        return chart

    def line_chart(ws, header_row: int, last_row: int, title: str, num_series: int = 2):
        chart = LineChart()
        chart.style = 10
        chart.title = title
        for col in range(2, 2 + num_series):
            data = Reference(ws, min_col=col, min_row=header_row, max_col=col, max_row=last_row)
            chart.add_data(data, titles_from_data=True)
        cats = Reference(ws, min_col=1, min_row=header_row + 1, max_row=last_row)
        chart.set_categories(cats)
        return chart

    # ── نمودار ۴-۱: Inertia + Silhouette ──
    ce = DATA["clustering_evaluation"]
    ws = wb.create_sheet("نمودار۴-۱")
    hr = write_title(ws, ce["title"])
    headers = ["k", "Inertia", "Silhouette"]
    rows = list(zip(ce["k"], ce["inertia"], ce["silhouette"]))
    write_headers(ws, hr, headers)
    last = write_rows(ws, hr + 1, rows)
    place_chart(ws, line_chart(ws, hr, last, "Inertia & Silhouette", 2), f"E{hr}")

    # ── نمودار ۴-۲: Inertia ──
    ws = wb.create_sheet("نمودار۴-۲")
    hr = write_title(ws, "نمودار ۴-۲ — تغییرات Inertia در مقادیر مختلف k")
    write_headers(ws, hr, ["k", "Inertia"])
    last = write_rows(ws, hr + 1, list(zip(ce["k"], ce["inertia"])))
    place_chart(ws, col_chart(ws, hr, last, "Inertia (WCSS)", "Inertia"), f"D{hr}")

    # ── نمودار ۴-۳: Silhouette ──
    ws = wb.create_sheet("نمودار۴-۳")
    hr = write_title(ws, "نمودار ۴-۳ — تغییرات ضریب Silhouette در مقادیر مختلف k")
    write_headers(ws, hr, ["k", "Silhouette"])
    last = write_rows(ws, hr + 1, list(zip(ce["k"], ce["silhouette"])))
    place_chart(ws, col_chart(ws, hr, last, "Silhouette", "ضریب"), f"D{hr}")

    # ── نمودار ۴-۴: اندازه خوشه‌ها ──
    cs = DATA["cluster_sizes"]
    ws = wb.create_sheet("نمودار۴-۴")
    hr = write_title(ws, cs["title"])
    write_headers(ws, hr, ["خوشه", "تعداد اعضا"])
    last = write_rows(ws, hr + 1, list(zip([clean_label(l) for l in cs["labels"]], cs["sizes"])))
    place_chart(ws, col_chart(ws, hr, last, "تعداد اعضای خوشه"), f"D{hr}")

    # ── نمودار ۴-۵: وزن AHP ──
    aw = DATA["ahp_weights"]
    ws = wb.create_sheet("نمودار۴-۵")
    hr = write_title(ws, aw["title"])
    write_headers(ws, hr, ["معیار", "سهم (%)"])
    last = write_rows(ws, hr + 1, list(zip(aw["criteria"], aw["percent"])))
    place_chart(ws, bar_h_chart(ws, hr, last, "وزن معیارها (AHP)", "سهم (%)"), f"D{hr}")

    # ── نمودار ۴-۶: وزن پایه vs تطبیقی ──
    wc = DATA["weight_comparison"]
    ws = wb.create_sheet("نمودار۴-۶")
    hr = write_title(ws, wc["title"])
    write_headers(ws, hr, ["معیار", "وزن پایه AHP", "وزن تطبیقی"])
    last = write_rows(ws, hr + 1, list(zip(wc["criteria"], wc["base"], wc["adaptive"])))
    place_chart(ws, col_chart(ws, hr, last, "مقایسه وزن‌ها", "وزن", data_cols=2), f"E{hr}")

    # ── نمودار ۴-۷: وزن تطبیقی ──
    ad = DATA["adaptive_weights"]
    ws = wb.create_sheet("نمودار۴-۷")
    hr = write_title(ws, ad["title"])
    pct = [round(w * 100, 2) for w in ad["weights"]]
    write_headers(ws, hr, ["معیار", "سهم (%)"])
    last = write_rows(ws, hr + 1, list(zip(ad["criteria"], pct)))
    place_chart(ws, bar_h_chart(ws, hr, last, "وزن تطبیقی", "سهم (%)"), f"D{hr}")

    # ── نمودار ۴-۸: TOPSIS ──
    tp = DATA["topsis"]
    ws = wb.create_sheet("نمودار۴-۸")
    hr = write_title(ws, tp["title"])
    write_headers(ws, hr, ["فناوری", "ضریب نزدیکی (C_i)"])
    last = write_rows(ws, hr + 1, list(zip(techs, tp["closeness"])))
    place_chart(ws, col_chart(ws, hr, last, "TOPSIS — ضریب نزدیکی"), f"D{hr}")

    # ── نمودار ۴-۹: VIKOR (S, R, Q) ──
    vk = DATA["vikor"]
    ws = wb.create_sheet("نمودار۴-۹")
    hr = write_title(ws, vk["title_indices"])
    write_headers(ws, hr, ["فناوری", "S_i", "R_i", "Q_i"])
    last = write_rows(ws, hr + 1, list(zip(techs, vk["s"], vk["r"], vk["q"])))
    place_chart(ws, col_chart(ws, hr, last, "VIKOR — شاخص‌ها", data_cols=3), f"F{hr}")

    # ── نمودار ۴-۹ب: VIKOR Q ──
    ws = wb.create_sheet("نمودار۴-۹ب")
    hr = write_title(ws, vk["title_q"])
    write_headers(ws, hr, ["فناوری", "Q_i"])
    last = write_rows(ws, hr + 1, list(zip(techs, vk["q"])))
    place_chart(ws, col_chart(ws, hr, last, "VIKOR — شاخص نهایی Q"), f"D{hr}")

    # ── نمودار ۴-۱۰: COPRAS ──
    cp = DATA["copras"]
    ws = wb.create_sheet("نمودار۴-۱۰")
    hr = write_title(ws, cp["title"])
    write_headers(ws, hr, ["فناوری", "Q_i", "N_i (%)"])
    last = write_rows(ws, hr + 1, list(zip(techs, cp["q"], cp["n"])))
    place_chart(ws, col_chart(ws, hr, last, "COPRAS", data_cols=2), f"E{hr}")

    # ── نمودار ۴-۱۱: مقایسه رتبه ──
    rc = DATA["ranking_comparison"]
    ws = wb.create_sheet("نمودار۴-۱۱")
    hr = write_title(ws, rc["title"])
    write_headers(ws, hr, ["فناوری", "TOPSIS", "VIKOR", "COPRAS"])
    last = write_rows(ws, hr + 1, list(zip(techs, rc["topsis"], rc["vikor"], rc["copras"])))
    place_chart(ws, col_chart(ws, hr, last, "مقایسه رتبه", data_cols=3), f"F{hr}")

    # ── نمودار ۴-۱۲: اسپیرمن ──
    sp = DATA["spearman"]
    ws = wb.create_sheet("نمودار۴-۱۲")
    hr = write_title(ws, sp["title"])
    write_headers(ws, hr, ["جفت روش‌ها", "ضریب اسپیرمن"])
    last = write_rows(ws, hr + 1, list(zip([clean_label(p) for p in sp["pairs"]], sp["rho"])))
    place_chart(ws, col_chart(ws, hr, last, "همبستگی اسپیرمن"), f"D{hr}")

    # تنظیم عرض ستون‌ها
    for ws in wb.worksheets:
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 16

    path = OUTPUT_DIR / "chapter4-data.xlsx"
    wb.save(path)
    print(f"  Excel: {path.name} ({len(wb.worksheets)} sheets with charts)")


def main():
    global DATA, USE_FA_DIGITS
    DATA = load_data()
    USE_FA_DIGITS = DATA.get("settings", {}).get("use_persian_digits", True)

    print(f"Data: {DATA_FILE}")
    print(f"Output: {OUTPUT_DIR}\n")
    print("Generating figures...")

    fig_clustering_inertia_silhouette()
    fig_clustering_inertia_only()
    fig_clustering_silhouette_only()
    fig_cluster_sizes()
    fig_ahp_weights()
    fig_base_vs_adaptive_weights()
    fig_adaptive_weights_only()
    fig_topsis_closeness()
    fig_vikor_indices()
    fig_vikor_q_only()
    fig_copras_results()
    fig_ranking_comparison()
    fig_spearman_correlation()

    if DATA.get("settings", {}).get("export_excel", True):
        export_excel()

    png_count = len(list(OUTPUT_DIR.glob("*.png")))
    svg_count = len(list(OUTPUT_DIR.glob("*.svg")))
    print(f"\nDone: {png_count} PNG, {svg_count} SVG")


if __name__ == "__main__":
    main()
